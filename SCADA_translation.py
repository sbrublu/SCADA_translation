### Import libraries

import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import re
import asyncio
from googletrans import Translator
import nltk
nltk.download('wordnet')
from nltk.corpus import wordnet
from tqdm import tqdm
import openpyxl
import os
import shutil

### Functions

## Load excel file and create dataframe
# Handle window close event
def on_window_close(root):
    print("\nOperation canceled. Exiting...")
    root.quit()
    exit()

# Name selection
def on_confirm_main(listbox, selected_name, names, root):
    selected_index = listbox.curselection()
    if selected_index:
        selected_name[0] = names[selected_index[0]]
    else:
        messagebox.showinfo("Warning", "Please select a proper name to confirm")
    root.quit()

# Display widget for selecting a name on a list, widget is looped till selection is made
def select_name(names, message):
    # Initialize selection
    selected_name = [None]  # Use a list to store the selected name

    while selected_name[0] is None:
        root = tk.Tk()
        root.title(f"Select the {message}")

        # Handle window close event
        root.protocol("WM_DELETE_WINDOW", lambda: on_window_close(root))

        # Dialog window features
        root.geometry("400x200")  # Set a minimum size for the window, min 200 to see confirm button
        frame = tk.Frame(root)
        frame.pack(fill=tk.BOTH, expand=True)  # Ensure the frame is packed
        listbox = tk.Listbox(frame, selectmode=tk.SINGLE)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # List of options
        for name in names:
            listbox.insert(tk.END, name)

        #  Confirm button
        tk.Button(root, text="Confirm", command=lambda: on_confirm_main(listbox, selected_name, names, root)).pack()
        root.mainloop()
        root.destroy()

    return selected_name[0]

# Select excel file, sheet and colum, extract dataframe
def load_file():
    messagebox.showinfo("Close", "If the excel file you want to process is open, please close it before proceeding")
    file_name = filedialog.askopenfilename(title="Load the excel file you want to process", filetypes=[("Excel files", "*.xls;*.xlsx;*.xlsm")])
    if not file_name:
        print("No file selected. Exiting...")
        exit()

    excel_file = pd.ExcelFile(file_name)
    sheets = excel_file.sheet_names
    sheet_name = select_name(sheets, "excel sheet")

    return file_name, sheet_name

# Load the selected columns from the specified sheet in the excel file
def load_columns(file_name, sheet_name):
    excel_file = pd.ExcelFile(file_name)
    # Load the selected excel sheet into a dataframe
    try:
        cols = excel_file.parse(sheet_name).columns.tolist()
        src_col = select_name(cols, "source column")
        trans_col = select_name(cols, "translated coulmn")

    except Exception as e:
        print(f"\nError: {str(e)}. Exiting...")
        exit()

    return src_col, trans_col

# Load the specified columns from the excel file into a dataframe
def load_df(file_name, sheet_name, src_col, trans_col):
    # Load the selected excel sheet into a dataframe
    try:
        file_df = pd.read_excel(file_name, sheet_name=sheet_name, usecols=[src_col, trans_col])
        print(f"\n{file_df.fillna("").to_string(max_rows=20)}")  # Replace nan with empty string

    except Exception as e:
        print(f"\nError: {str(e)}. Exiting...")
        exit()

    return file_df

## Translate col_origin to col_translated
# Get a synonym for a word using WordNet
def get_synonym(word):
    # Try to get a synonym using WordNet
    synsets = wordnet.synsets(word)
    for syn in synsets:
        for lemma in syn.lemmas():
            synonym = lemma.name().replace('_', ' ')
            if synonym.lower() != word.lower():
                return synonym
    return word

# Abbreviate a word
# Shorten the translation using synonyms or abbreviations
def shorten_translation(original, translated, delta):
    words = translated.split()
    # 1. Try to shorten with synonyms
    new_words = []
    for w in words:
        synonym = get_synonym(w)
        if len(synonym) < len(w):
            new_words.append(synonym)
        else:
            new_words.append(w)
    # 2. If still too long, abbreviate longest words as needed using replace
    word_forms = [(w, w) for w in new_words]
    sorted_indices = sorted(range(len(new_words)), key=lambda i: -len(new_words[i]))
    shortened = translated
    for idx in sorted_indices:
        orig, _ = word_forms[idx]
        if len(shortened) <= len(original) or delta <= 0 or len(orig) <= 5:
            break
        abbr_len = max(4, len(orig) - delta - 1)
        abbr = orig[:abbr_len] + '.'
        delta -= (len(orig) - len(abbr))
        word_forms[idx] = (orig, abbr)
        # Replace only the first occurrence of the word in the string
        shortened = re.sub(r'\b{}\b'.format(re.escape(orig)), abbr, shortened, count=1)
        # Adjust delta based on the effective length change
    return shortened

# Apply the translation mapping with fallback to Google Translate
async def translate_value(value, dictionary, src_lang, trans_lang):
    translator = Translator()
    value_key = str(value).strip().lower()

    # Check if the value is in the translation map
    if value in dictionary and pd.notna(dictionary[value]):
        translation = dictionary[value]  # Use dictionary translation
        class Dummy: pass
        dummy = Dummy()
        dummy.text = translation
        translation = dummy
    else:
        # If not in the map, use Google Translate
        try:
            translation = await translator.translate(value, src=src_lang, dest=trans_lang)

        except Exception as e:
            print(f"Translation error for '{value}': {str(e)}")
            return value, value  # Return the original value if translation fails

    # Preserve the formatting of the original string
    if value.isupper():
        translated = translation.text.upper()
    elif value.istitle():
        translated = translation.text.title()
    elif value.islower():
        translated = translation.text.lower()
    else:
        translated = translation.text

    # Enforce length limit
    delta = len(translated) - len(value)
    if delta > 0:
        translated_short = shorten_translation(value, translated, delta)
    else:
        translated_short = translated

    return translated, translated_short

# Translate the specified column in the dataframe
async def translate_column_async(trans_df, dict_df, src_col, trans_col, src_lang, trans_lang):
    # Extract unique values containing alphabetic words
    word_pattern = re.compile(r'\b[a-zA-Z]+\b')
    unique_strings = {
        match.strip()
        for val, trans_val in zip(trans_df[src_col], trans_df[trans_col])
        if pd.isna(trans_val) and isinstance(val, str)
        for match in re.findall(r'(?:\b[a-zA-Z]+\b(?:\s+)?)+', val)
        if any(word.isalpha() for word in word_pattern.findall(val))
    }

    dictionary = {
        str(row[src_lang]).strip().lower(): row[trans_lang]
        for _, row in dict_df.iterrows()
        if pd.notna(row[src_lang]) and pd.notna(row[trans_lang])
    }

    # Translate unique values
    translation_map = {
        val: await translate_value(val, dictionary, src_lang, trans_lang)
        for val in tqdm(unique_strings, desc="Translating unique values")
    }

    # Print translations
    print("\nOriginal -> Translated -> Shortened")
    for val, (translated, translated_short) in translation_map.items():
        print(f"{val} -> {translated} -> {translated_short}")

    # Build the list of dictionaries for the DataFrame
    data = [{src_lang: val, trans_lang: translated_short} for val, (translated, translated_short) in translation_map.items()]

    # Create the DataFrame
    dict_df_new = pd.DataFrame(data, columns=[src_lang, trans_lang])

    print(f"\n{dict_df_new.fillna('').to_string(max_rows=20)}")

    # Apply shortened translations to the dataframe
    trans_df[trans_col] = trans_df.apply(
        lambda row: translation_map[row[src_col]][1] if (
                    pd.isna(row[trans_col]) and row[src_col] in translation_map
        ) else (row[src_col] if pd.isna(row[trans_col]) else row[trans_col]),
        axis=1
    )

    print(f"\n{trans_df.fillna('').to_string(max_rows=20)}")

    return trans_df, dict_df_new

## Write to excel file
# Copy the original file to a new file with "_translated" suffix
def copy_file(file_name):
    try:
        # Create a copy of the original file
        base, ext = os.path.splitext(file_name)
        copied_file_name = base + "_translated" + ext
        shutil.copy(file_name, copied_file_name)

    except Exception as e:
        print(f"\nError copying the file: {str(e)}. Exiting...")
        exit()

    return copied_file_name

# Write translated column to excel file
def write_col(file_name, sheet_name, file_df, col, start_row):
    try:
        base, ext = os.path.splitext(file_name)
        # Use keep_vba only for .xlsm files
        if ext.lower() == ".xlsm":
            wb = openpyxl.load_workbook(file_name, keep_vba=True)
        else:
            wb = openpyxl.load_workbook(file_name)

        if sheet_name not in wb.sheetnames:
            print(f"\nSheet {sheet_name} not found in the workbook.Exiting...")
            exit()

        ws = wb[sheet_name]
        header = [cell.value for cell in ws[1]]
        col_idx = header.index(col) + 1
        row_idx = 2

        if start_row != 2:
            # Find the first empty row in src_col
            while ws.cell(row=row_idx, column=col_idx).value not in (None, ""):
                row_idx += 1
            start_row = row_idx

        # Write the translated column to the sheet
        for r_idx, value in enumerate(file_df[col], start=start_row):
            ws.cell(row=r_idx, column=col_idx, value=value)

        wb.save(file_name)
        wb.close()

    except Exception as e:
        print(f"\nError writing to Excel: {str(e)}. Exiting...")
        exit()

    return start_row

### Main

if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()

    # Load excel file and create dataframes for map and list
    messagebox.showinfo("Empty", "Loading file to translate. Words present in the translated columns will be kept. Please make sure to delete unwanted cells before proceeding")
    trans_file, trans_sheet = load_file()
    trans_src, trans_trans = load_columns(trans_file, trans_sheet)
    trans_df = load_df(trans_file, trans_sheet, trans_src, trans_trans)

    # Select languages for translation
    languages = ["en", "es", "fr", "de", "it", "pt"]
    src_lang = select_name(languages, "source language")
    trans_lang = select_name(languages, "translated language")

    # Load excel file for dictionary
    messagebox.showinfo("Empty", "Loading dictionary file")
    dict_file, dict_sheet = load_file()
    dict_df = load_df(dict_file, dict_sheet, src_lang, trans_lang)

    # Run the translation
    trans_df, dict_df_new = asyncio.run(translate_column_async(trans_df, dict_df, trans_src, trans_trans, src_lang, trans_lang))

    proceed_write_trans = messagebox.askyesno("Confirmation", "Do you want to write translation to the excel file?")
    if proceed_write_trans:
        # Write the translated column to the excel file
        trans_copy = copy_file(trans_file)
        start_row = write_col(trans_copy, trans_sheet, trans_df, trans_trans, 2)

    proceed_write_dict = messagebox.askyesno("Confirmation", "Do you want to write dictionary to the excel file?")
    if not proceed_write_dict:
        print("\nOperation canceled. Exiting...")
        exit()

    # Load the dictionary file and write the dictionary dataframe
    dict_copy = copy_file(dict_file)
    start_row = write_col(dict_copy, dict_sheet, dict_df_new, src_lang, 0)
    start_row = write_col(dict_copy, dict_sheet, dict_df_new, trans_lang, start_row)